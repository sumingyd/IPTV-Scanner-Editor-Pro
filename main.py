from PyQt6 import QtWidgets, QtCore, QtGui
import time
import threading
from channel_model import ChannelListModel
from epg_manager import EPGManager
from ui_builder import UIBuilder
from config_manager import ConfigManager
from log_manager import LogManager
from scanner_controller import ScannerController
from styles import AppStyles
from epg_ui import EPGManagementDialog, EPGProgramWidget
import sys
import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Dict

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        # 初始化配置和日志管理器
        self.config = ConfigManager()
        self.logger = LogManager()
        
        # 初始化频道映射
        from channel_mappings import REVERSE_MAPPINGS
        self.channel_mappings = REVERSE_MAPPINGS
        
        # 初始化EPG管理器
        from epg_manager import EPGManager
        self.epg_manager = EPGManager(self.config, self)
        
        # 构建UI
        self.ui = UIBuilder(self)
        self.ui.build_ui()
        
        # 初始化控制器(确保模型已由UIBuilder初始化)
        self.scanner = ScannerController(self.model, self.epg_manager)
        from player_controller import PlayerController
        from list_manager import ListManager
        self.player_controller = PlayerController(self.ui.main_window.player, self.model)
        self.list_manager = ListManager(self.model)
        
        # UI构建完成后加载配置
        self._load_config()
        
        # 初始化频道名称自动补全
        self._setup_name_autocomplete()
        
        # 连接信号槽
        self._connect_signals()

    def _update_validate_status(self, message):
        """更新有效性检测状态标签"""
        self.ui.main_window.validate_stats_label.setText(message)
        message == "请点击检测有效性按钮"

    def _setup_name_autocomplete(self):
        """设置频道名称自动补全"""
        completer = QtWidgets.QCompleter()
        completer.setCaseSensitivity(QtCore.Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(QtCore.Qt.MatchFlag.MatchContains)
        model = QtCore.QStringListModel()
        completer.setModel(model)
        self.ui.main_window.name_edit.setCompleter(completer)

    def _load_config(self):
        """加载保存的配置到UI"""
        try:
            settings = self.config.load_network_settings()
            if settings['url']:
                self.ui.main_window.ip_range_input.setText(settings['url'])
            self.ui.main_window.timeout_input.setValue(int(settings['timeout']))
            self.ui.main_window.thread_count_input.setValue(int(settings['threads']))
            if settings['user_agent']:
                self.ui.main_window.user_agent_input.setText(settings['user_agent'])
            if settings['referer']:
                self.ui.main_window.referer_input.setText(settings['referer'])
        except Exception as e:
            self.logger.error(f"加载网络设置失败: {e}")
            # 设置默认值
            self.ui.main_window.timeout_input.setValue(10)
            self.ui.main_window.thread_count_input.setValue(5)

    def _connect_signals(self):
        """连接所有信号和槽"""
        # 连接频道列表选择信号
        self.ui.main_window.channel_list.selectionModel().selectionChanged.connect(
            self._on_channel_selected
        )
        
        # 连接扫描按钮
        try:
            self.ui.main_window.scan_btn.clicked.disconnect()
        except:
            pass
        self.ui.main_window.scan_btn.clicked.connect(self._on_scan_clicked)
        
        # 连接菜单栏动作
        for action in self.ui.main_window.menuBar().actions():
            try:
                action.triggered.disconnect()
            except:
                pass
            if action.text().startswith("打开列表"):
                action.triggered.connect(self._open_list)
            elif action.text().startswith("保存列表"):
                action.triggered.connect(self._save_list)
                
        # 连接工具栏按钮
        for action in self.ui.main_window.findChildren(QtGui.QAction):
            try:
                action.triggered.disconnect()
            except:
                pass
            if "打开列表" in action.text():
                action.triggered.connect(self._open_list)
            elif "保存列表" in action.text():
                action.triggered.connect(self._save_list)
            elif "导入Excel" in action.text():
                action.triggered.connect(self.ui._import_excel)
            elif "导出Excel" in action.text():
                action.triggered.connect(self.ui._export_excel)
            elif "刷新EPG" in action.text():
                action.triggered.connect(self._on_refresh_epg_clicked)
            elif "EPG管理" in action.text():
                action.triggered.connect(self._on_epg_manager_clicked)
            elif "关于" in action.text():
                action.triggered.connect(self._on_about_clicked)
                
        # 连接播放控制信号
        self.ui.main_window.volume_slider.valueChanged.connect(
            self._on_volume_changed)
        self.ui.main_window.pause_btn.clicked.connect(
            self._on_pause_clicked)
        self.ui.main_window.stop_btn.clicked.connect(
            self._on_stop_clicked)
        
        # 连接播放状态变化信号
        self.player_controller.play_state_changed.connect(
            self._on_play_state_changed)
        
        # 连接频道列表双击事件
        self.ui.main_window.channel_list.doubleClicked.connect(self._play_selected_channel)
        
        # 连接有效性检测按钮
        self.ui.main_window.btn_validate.clicked.connect(self._on_validate_clicked)
        
        # 连接隐藏无效项按钮
        self.ui.main_window.btn_hide_invalid.clicked.connect(self._on_hide_invalid_clicked)
        
        # 进度更新信号
        self.scanner.progress_updated.connect(
            lambda cur, total: self.ui.main_window.scan_progress.setValue(
                int(cur / total * 100) if total > 0 else 0
            )
        )
        
        # 频道发现信号
        self.scanner.channel_found.connect(self._on_channel_found)
        
        # 扫描完成信号
        self.scanner.scan_completed.connect(self._on_scan_completed)
        
        # 统计信息更新信号
        self.scanner.stats_updated.connect(self._update_stats_display)

    def _on_scan_clicked(self):
        """处理扫描按钮点击事件"""
        if self.scanner.is_scanning():
            # 停止扫描
            self.scanner.stop_scan()
            self.ui.main_window.scan_btn.setText("完整扫描")
        else:
            # 检查地址是否为空
            url = self.ui.main_window.ip_range_input.text()
            if not url.strip():
                self.logger.warning("请输入扫描地址")
                self.ui.main_window.statusBar().showMessage("请输入扫描地址", 3000)
                return
                
            # 开始扫描前清空列表
            self.model.clear()
            timeout = self.ui.main_window.timeout_input.value()
            threads = self.ui.main_window.thread_count_input.value()
            self.scanner.start_scan(url, threads, timeout)
            self.ui.main_window.scan_btn.setText("停止扫描")

    def _validate_all_channels(self, timeout: int, threads: int):
        """验证所有频道的有效性"""
        self.scanner.timeout = timeout
        self.scanner.stop_event.clear()
        
        # 初始化统计信息
        self.scanner.stats = {
            'total': self.ui.main_window.model.rowCount(),
            'valid': 0,
            'invalid': 0,
            'start_time': time.time(),
            'elapsed': 0
        }
        
        # 填充任务队列
        for i in range(self.ui.main_window.model.rowCount()):
            channel = self.ui.main_window.model.get_channel(i)
            self.scanner.worker_queue.put(channel['url'])
            
        # 创建工作线程
        self.scanner.workers = []
        for i in range(threads):
            worker = threading.Thread(
                target=self.scanner._worker,
                name=f"ValidatorWorker-{i}",
                daemon=True
            )
            worker.start()
            self.scanner.workers.append(worker)
            
        # 启动统计更新线程
        stats_thread = threading.Thread(
            target=self.scanner._update_stats,
            name="StatsUpdater",
            daemon=True
        )
        stats_thread.start()

    def _on_volume_changed(self, value):
        """处理音量滑块变化"""
        self.player_controller.set_volume(value)

    def _on_pause_clicked(self):
        """处理暂停/播放按钮点击"""
        is_playing = self.player_controller.toggle_pause()
        self.ui.main_window.pause_btn.setText("暂停" if is_playing else "播放")

    def _on_stop_clicked(self):
        """处理停止按钮点击"""
        self.player_controller.stop()
        self.ui.main_window.pause_btn.setText("播放")

    def _on_play_state_changed(self, is_playing):
        """处理播放状态变化"""
        self.ui.main_window.stop_btn.setEnabled(is_playing)
        self.ui.main_window.stop_btn.setStyleSheet(
            AppStyles.button_style(active=is_playing)
        )

    def _open_list(self):
        """打开列表文件"""
        try:
            # 获取打开文件路径
            file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
                self,
                "打开频道列表",
                "",
                "Excel文件 (*.xlsx);;文本文件 (*.txt);;所有文件 (*)"
            )
            
            if not file_path:
                return False
                
            # 根据文件扩展名选择打开方式
            if file_path.lower().endswith('.xlsx'):
                # 读取Excel文件
                with open(file_path, 'rb') as f:
                    excel_data = f.read()
                result = self._handle_excel_import(excel_data)
            else:
                # 打开文本文件
                result = self.list_manager.open_list(self, file_path)
                
            if result:
                self.ui.main_window.btn_hide_invalid.setEnabled(False)
                self.ui.main_window.btn_validate.setEnabled(True)
                self.ui.main_window.statusBar().showMessage("列表加载成功", 3000)
                
                # 强制触发模型重置信号
                self.model.modelReset.emit()
                # 立即更新按钮状态
                self.ui._update_load_button_state()
                # 重置智能匹配区域状态
                self.ui.match_status_label.setText("智能匹配状态: 等待操作")
                self.ui.main_window.match_progress.setValue(0)
                return True
            else:
                self.logger.warning("打开列表失败")
                self.ui.main_window.statusBar().showMessage("打开列表失败", 3000)
                return False
        except Exception as e:
            self.logger.error(f"打开列表失败: {str(e)}", exc_info=True)
            self.ui.main_window.statusBar().showMessage(f"打开列表失败: {str(e)}", 3000)
            return False

    def load_old_list(self, file_path):
        """加载旧列表文件到内存"""
        try:
            channels = self.list_manager.load_old_list(file_path)
            if channels:
                # 将频道数据存储在内存中，供后续使用
                self.old_channels = channels
                return True
            else:
                self.logger.warning("旧列表加载失败")
                return False
        except Exception as e:
            self.logger.error(f"加载旧列表失败: {str(e)}", exc_info=True)
            return False

    def _handle_excel_import(self, excel_data: bytes) -> bool:
        """处理Excel导入数据"""
        try:
            # 创建临时文件
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp.write(excel_data)
                tmp_path = tmp.name
            
            # 加载Excel文件
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active
            
            # 读取表头
            headers = [cell.value for cell in ws[1]]
            
            # 确定各字段位置
            col_map = {}
            for idx, header in enumerate(headers, 1):
                if header:
                    header_lower = str(header).lower()
                    if 'name' in header_lower:
                        col_map['name'] = idx
                    elif 'url' in header_lower:
                        col_map['url'] = idx
                    elif 'group' in header_lower:
                        col_map['group'] = idx
                    elif 'logo' in header_lower:
                        col_map['logo'] = idx
                    elif 'valid' in header_lower:
                        col_map['valid'] = idx
                    elif 'delay' in header_lower:
                        col_map['delay'] = idx
            
            # 检查必要字段
            if 'name' not in col_map or 'url' not in col_map:
                self.logger.error("Excel缺少必要列: 名称或URL")
                return False
            
            # 读取数据行
            channels = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # 跳过空行
                    continue
                
                channel = {
                    'name': str(row[col_map['name']-1]) if col_map.get('name') else '',
                    'url': str(row[col_map['url']-1]) if col_map.get('url') else '',
                    'group': str(row[col_map.get('group', 0)-1]) if col_map.get('group') else '未分类',
                    'logo': str(row[col_map.get('logo', 0)-1]) if col_map.get('logo') else None,
                    'valid': bool(row[col_map.get('valid', 0)-1]) if col_map.get('valid') else False,
                    'delay': float(row[col_map.get('delay', 0)-1]) if col_map.get('delay') else 0.0
                }
                channels.append(channel)
            
            # 更新频道列表
            self.model.clear()
            for channel in channels:
                self.model.add_channel(channel)
            
            # 删除临时文件
            import os
            os.unlink(tmp_path)
            
            self.logger.info(f"成功导入 {len(channels)} 个频道")
            return True
            
        except Exception as e:
            self.logger.error(f"导入Excel失败: {str(e)}", exc_info=True)
            return False

    def _generate_excel_data(self) -> bytes:
        """生成Excel格式的频道数据"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # 写入表头
            headers = [
                "频道名称", "URL", "分组", "Logo路径", 
                "有效性", "延迟(秒)", "分辨率", "状态"
            ]
            ws.append(headers)
            
            # 写入数据行
            for i in range(self.model.rowCount()):
                channel = self.model.get_channel(i)
                from channel_mappings import get_channel_info
                channel_info = get_channel_info(channel.get('name', ''))
                ws.append([
                    channel.get('name', ''),
                    channel.get('url', ''),
                    channel.get('group', '未分类'),
                    channel_info.get('logo_url') or channel.get('logo', ''),
                    channel.get('valid', False),
                    channel.get('latency', 0.0) if channel.get('latency') else 0.0,
                    channel.get('resolution', ''),
                    channel.get('status', '')
                ])
            
            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            self.logger.error(f"生成Excel数据失败: {str(e)}", exc_info=True)
            raise

    def _save_list(self):
        """保存列表文件"""
        try:
            # 获取保存文件路径
            file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
                self,
                "保存频道列表",
                "",
                "Excel文件 (*.xlsx);;文本文件 (*.txt);;所有文件 (*)"
            )
            
            if not file_path:
                return False
                
            # 根据文件扩展名选择保存方式
            if file_path.lower().endswith('.xlsx'):
                # 保存为Excel格式
                excel_data = self._generate_excel_data()
                with open(file_path, 'wb') as f:
                    f.write(excel_data)
                self.logger.info("Excel列表保存成功")
                self.ui.main_window.statusBar().showMessage("Excel列表保存成功", 3000)
                return True
            else:
                # 保存为文本格式
                result = self.list_manager.save_list(self, file_path)
                if result:
                    self.logger.info("列表保存成功")
                    self.ui.main_window.statusBar().showMessage("列表保存成功", 3000)
                    return True
                else:
                    self.logger.warning("列表保存失败")
                    self.ui.main_window.statusBar().showMessage("列表保存失败", 3000)
                    return False
                    
        except Exception as e:
            self.logger.error(f"保存列表失败: {e}", exc_info=True)
            self.ui.main_window.statusBar().showMessage(f"保存列表失败: {str(e)}", 3000)
            return False

    def _on_validate_clicked(self):
        """处理有效性检测按钮点击事件"""
        if not self.ui.main_window.model.rowCount():
            self.logger.warning("请先加载列表")
            return
            
        if not hasattr(self.scanner, 'is_validating') or not self.scanner.is_validating:
            # 开始有效性检测
            timeout = self.ui.main_window.timeout_input.value()
            threads = self.ui.main_window.thread_count_input.value()
            self.scanner.start_validation(
                self.ui.main_window.model,
                threads,
                timeout
            )
            self.ui.main_window.btn_validate.setText("停止检测")
            self.ui.main_window.btn_hide_invalid.setEnabled(True)
            self.ui.main_window.btn_hide_invalid.setStyleSheet(
                AppStyles.button_style(active=True)
            )
            
            # 连接验证结果信号
            self.scanner.channel_validated.connect(self._on_channel_validated)
        else:
            # 停止有效性检测
            self.scanner.stop_validation()
            self.ui.main_window.btn_validate.setText("检测有效性")
            
    def _on_channel_selected(self):
        """处理频道选择事件"""
        selected = self.ui.main_window.channel_list.selectedIndexes()
        if not selected:
            return
            
        # 获取选中的频道
        row = selected[0].row()
        self.current_channel_index = row
        channel = self.model.get_channel(row)
        
        # 更新编辑框
        self.ui.main_window.name_edit.setText(channel.get('name', ''))
        self.ui.main_window.group_combo.setCurrentText(channel.get('group', '未分类'))
        
        # 直接访问保存按钮对象
        self.ui.main_window.save_channel_btn.setEnabled(True)
        try:
            self.ui.main_window.save_channel_btn.clicked.disconnect()  # 先断开所有连接
        except:
            pass
        self.ui.main_window.save_channel_btn.clicked.connect(self._on_save_clicked)

    def _on_save_clicked(self):
        """处理保存按钮点击事件"""
        try:
            # 检查当前选中频道
            if not hasattr(self, 'current_channel_index'):
                self.logger.warning("保存失败: 未选中频道")
                self.ui.main_window.statusBar().showMessage("请先选择要编辑的频道", 3000)
                return
                
            # 检查模型是否有效
            if not hasattr(self, 'model') or not self.model:
                self.logger.error("保存失败: 频道模型未初始化")
                self.ui.main_window.statusBar().showMessage("系统错误: 频道模型未初始化", 3000)
                return
                
            # 检查EPG组件状态
            epg_status = "正常"
            if not hasattr(self, 'epg_manager'):
                epg_status = "EPG管理器未初始化"
            elif not hasattr(self.epg_manager, 'is_loaded'):
                epg_status = "EPG数据未加载"
            self.logger.debug(f"EPG状态: {epg_status}")
            
            # 获取并验证编辑数据
            name = self.ui.main_window.name_edit.text().strip()
            group = self.ui.main_window.group_combo.currentText().strip()
            
            if not name:
                self.logger.warning("保存失败: 频道名不能为空")
                self.ui.main_window.statusBar().showMessage("频道名不能为空", 3000)
                return
                
            # 获取并验证频道数据
            try:
                channel = self.model.get_channel(self.current_channel_index)
                if not channel or not isinstance(channel, dict):
                    raise ValueError("无效的频道数据")
                    
                # 创建频道数据副本
                new_channel = channel.copy()
                new_channel['name'] = name
                new_channel['group'] = group
                
                # 验证新数据
                if not all(key in new_channel for key in ['name', 'group', 'url']):
                    raise ValueError("频道数据不完整")
                    
                # 安全更新频道数据
                try:
                    self.model.update_channel(self.current_channel_index, new_channel)
                    
                    # 通知模型更新
                    self.model.dataChanged.emit(
                        self.model.index(self.current_channel_index, 0),
                        self.model.index(self.current_channel_index, self.model.columnCount() - 1)
                    )
                    
                    # 更新自动补全数据
                    self._update_name_completer(self.model.get_all_channel_names())
                    
                    self.logger.info(f"成功保存频道修改: {name} (原名称: {channel.get('name', '无')})")
                    self.ui.main_window.statusBar().showMessage("保存成功", 2000)
                    
                except Exception as update_error:
                    self.logger.error(f"更新频道数据失败: {str(update_error)}", exc_info=True)
                    raise ValueError("更新频道数据时出错")
                    
            except Exception as channel_error:
                self.logger.error(f"处理频道数据时出错: {str(channel_error)}", exc_info=True)
                self.ui.main_window.statusBar().showMessage(f"保存失败: {str(channel_error)}", 3000)
                return
                
        except Exception as e:
            self.logger.critical(f"保存操作发生严重错误: {str(e)}", exc_info=True)
            QtWidgets.QMessageBox.critical(
                self.ui.main_window,
                "保存错误",
                f"保存操作发生严重错误:\n{str(e)}\n请检查日志获取详细信息"
            )

    def _on_channel_validated(self, index, valid, latency, resolution):
        """处理频道验证结果"""
        channel = self.ui.main_window.model.get_channel(index)
        channel['valid'] = valid
        channel['latency'] = latency
        channel['resolution'] = resolution
        channel['status'] = '有效' if valid else '无效'
        
        # 通知模型更新
        self.ui.main_window.model.dataChanged.emit(
            self.ui.main_window.model.index(index, 0),
            self.ui.main_window.model.index(index, self.ui.main_window.model.columnCount() - 1)
        )

    def _on_hide_invalid_clicked(self):
        """处理隐藏无效项按钮点击事件"""
        if self.ui.main_window.btn_hide_invalid.text() == "隐藏无效项":
            self.ui.main_window.model.hide_invalid()
            self.ui.main_window.btn_hide_invalid.setText("恢复隐藏项")
        else:
            self.ui.main_window.model.show_all()
            self.ui.main_window.btn_hide_invalid.setText("隐藏无效项")

    def _play_selected_channel(self, index):
        """播放选中的频道"""
        if not index.isValid():
            return
            
        channel = self.ui.main_window.model.get_channel(index.row())
            
        if not hasattr(self, 'player_controller'):
            from player_controller import PlayerController
            self.player_controller = PlayerController(self.ui.main_window.player)
            
        if self.player_controller.play_channel(channel):
            self.ui.main_window.pause_btn.setText("暂停")
            self.current_channel = channel
            # 延迟执行确保播放器初始化完成
            QtCore.QTimer.singleShot(500, lambda: self._update_epg_display(channel))

    def _on_channel_found(self, channel_info):
        """处理发现有效频道事件"""
        self.ui.main_window.model.add_channel(channel_info)

    def _on_scan_completed(self):
        """处理扫描完成事件"""
        self.ui.main_window.scan_btn.setText("完整扫描")
        self.ui.main_window.btn_validate.setText("检测有效性")
        self.logger.info("扫描完成")
        # 扫描完成后强制更新"加载旧列表"按钮状态
        self.ui._update_load_button_state()
        # 重置智能匹配区域状态
        self.ui.match_status_label.setText("智能匹配状态: 等待操作")
        self.ui.main_window.match_progress.setValue(0)

    def _update_stats_display(self, stats_data):
        """更新统计信息显示"""
        if stats_data.get('is_validation', False):
            # 更新检测有效性统计标签
            self.ui.main_window.validate_stats_label.setText(stats_data['text'])
        else:
            # 更新扫描统计信息
            stats = stats_data.get('stats', {})
            elapsed = time.strftime("%H:%M:%S", time.gmtime(stats.get('elapsed', 0)))
            self.ui.main_window.detailed_stats_label.setText(
                f"总数: {stats.get('total', 0)} | "
                f"有效: {stats.get('valid', 0)} | "
                f"无效: {stats.get('invalid', 0)} | "
                f"耗时: {elapsed}"
            )

    def _on_refresh_epg_clicked(self):
        """处理刷新EPG按钮点击事件"""
        self.logger.info("用户点击刷新EPG按钮")
        try:
            # 检查是否按住Shift键(强制刷新)
            force_update = QtWidgets.QApplication.keyboardModifiers() == QtCore.Qt.KeyboardModifier.ShiftModifier
            
            # 显示进度指示器
            self.ui.main_window.progress_indicator.show()
            self.ui.main_window.epg_match_label.setText("EPG状态: 正在刷新...")
        finally:
            # 确保进度指示器被隐藏
            self.ui.main_window.progress_indicator.hide()
            
            # 创建信号对象用于线程间通信
            class RefreshSignals(QtCore.QObject):
                update_status = QtCore.pyqtSignal(str)
                update_completer = QtCore.pyqtSignal(list)
                update_epg_display = QtCore.pyqtSignal(dict)
                finished = QtCore.pyqtSignal()
                
            signals = RefreshSignals()
            
            # 连接信号槽
            signals.update_status.connect(self.ui.main_window.epg_match_label.setText)
            signals.update_completer.connect(self._update_name_completer)
            signals.update_epg_display.connect(lambda _: self._update_epg_display(self.current_channel))
            signals.finished.connect(self.ui.main_window.progress_indicator.hide)
            
            # 在后台线程中执行EPG刷新
            def refresh_task():
                try:
                    # 调用EPG刷新并等待完成
                    success = self.epg_manager.refresh_epg(force_update=force_update)
                    if not success:
                        self.logger.warning("EPG刷新启动失败")
                        signals.update_status.emit("EPG状态: 刷新启动失败")
                        return
                        
                    # 等待线程完成
                    while not self.epg_manager.isFinished():
                        QtCore.QCoreApplication.processEvents()
                        
                    # 检查EPG操作状态
                    if hasattr(self.epg_manager, 'last_operation_status'):
                        if self.epg_manager.last_operation_status:
                            self.logger.info("EPG操作成功")
                            signals.update_status.emit("EPG状态: 已加载")
                        else:
                            if not self.config.epg_sources:
                                self.logger.warning("EPG操作失败: 未配置EPG源")
                            else:
                                self.logger.warning("EPG操作失败: 本地无EPG数据")
                            if not self.config.epg_sources:
                                signals.update_status.emit("EPG状态: 未配置EPG源，请先添加EPG源")
                            else:
                                signals.update_status.emit("EPG状态: 本地无EPG数据，请先获取EPG数据")
                    else:
                        # 回退到检查result
                        if self.epg_manager.result:
                            self.logger.info("EPG刷新成功")
                            signals.update_status.emit("EPG状态: 已加载")
                        else:
                            self.logger.warning("EPG刷新失败")
                            signals.update_status.emit("EPG状态: 刷新失败")
                        
                    # 更新自动补全数据
                    channel_names = self.epg_manager.get_channel_names()
                    signals.update_completer.emit(channel_names)
                    # 更新当前播放频道的节目单
                    if hasattr(self, 'current_channel'):
                        signals.update_epg_display.emit(self.current_channel)
                except Exception as e:
                    self.logger.error(f"EPG刷新出错: {e}")
                    signals.update_status.emit("EPG状态: 刷新出错")
                finally:
                    signals.finished.emit()

            # 如果已有线程在运行，先停止它
            if hasattr(self, 'epg_thread') and self.epg_thread.isRunning():
                self.epg_thread.quit()
                self.epg_thread.wait(1000)
            
            # 创建新的线程和worker
            self.epg_thread = QtCore.QThread()
            self.epg_worker = QtCore.QObject()
            self.epg_worker.moveToThread(self.epg_thread)
            
            # 连接信号
            self.epg_thread.started.connect(refresh_task)
            
            # 线程结束时自动清理
            self.epg_thread.finished.connect(self.epg_thread.deleteLater)
            self.epg_thread.finished.connect(self.epg_worker.deleteLater)
            
            # 启动线程并确保资源释放
            try:
                self.epg_thread.start()
            except RuntimeError as e:
                self.logger.error(f"EPG线程启动失败: {e}")
                self.ui.main_window.epg_match_label.setText("EPG状态: 启动失败")
                self.epg_thread.quit()
                self.epg_thread.wait(1000)
                return
            except Exception as e:
                self.logger.error(f"EPG线程未知错误: {e}")
                self.ui.main_window.epg_match_label.setText("EPG状态: 未知错误")
                self.epg_thread.quit()
                self.epg_thread.wait(1000)
                return

    def closeEvent(self, event):
        if hasattr(self, 'epg_thread') and self.epg_thread and self.epg_thread.isRunning():
            self.epg_thread.quit()
            self.epg_thread.wait()
        event.accept()

    def _update_name_completer(self, channel_names):
        """更新频道名称自动补全数据"""
        completer = self.ui.main_window.name_edit.completer()
        if completer and channel_names:
            model = completer.model()
            if model:
                model.setStringList(channel_names)

    def _update_epg_display(self, channel):
        """更新EPG节目单显示"""
        if not hasattr(self, 'epg_manager') or not self.epg_manager:
            self.logger.warning("EPG管理器未初始化")
            self.ui.main_window.epg_title.setText(f"{channel['name']} - EPG未初始化")
            self.ui.main_window.epg_timeline.setWidget(QtWidgets.QLabel("EPG管理器未初始化"))
            return
            
        try:
            programs = self.epg_manager.get_channel_programs(channel['name'])
            
            if not programs:
                self.ui.main_window.epg_title.setText(f"{channel['name']} - 无节目数据")
                self.ui.main_window.epg_timeline.setWidget(QtWidgets.QLabel("没有可用的节目信息"))
                return
                
            self.logger.info(f"找到 {len(programs)} 个节目")
        except Exception as e:
            self.logger.error(f"获取节目数据出错: {e}")
            self.ui.main_window.epg_title.setText(f"{channel['name']} - 获取节目数据出错")
            self.ui.main_window.epg_timeline.setWidget(QtWidgets.QLabel(f"获取节目数据出错: {str(e)}"))
            return
        
        # 创建EPG节目单控件
        epg_widget = EPGProgramWidget()
        epg_widget.update_programs(channel['name'], programs)
        
        # 更新UI
        self.ui.main_window.epg_title.setText(f"{channel['name']} 节目单")
        self.ui.main_window.epg_content.setWidget(epg_widget)

    def _on_epg_manager_clicked(self):
        """处理EPG管理按钮点击事件"""
        dialog = EPGManagementDialog(self, self.config, self._save_epg_config)
        dialog.exec()

    def _on_about_clicked(self):
        """处理关于按钮点击事件"""
        try:
            from about_dialog import AboutDialog
            
            # 确保在主线程中创建对话框
            dialog = AboutDialog(self)
            dialog.setWindowModality(QtCore.Qt.WindowModality.ApplicationModal)
            
            # 确保事件循环正确处理
            def show_dialog():
                try:
                    dialog.show()
                except Exception as e:
                    self.logger.error(f"显示对话框出错: {e}")
            
            # 使用定时器确保UI更新
            QtCore.QTimer.singleShot(0, show_dialog)
            
        except ImportError as e:
            self.logger.error(f"导入AboutDialog失败: {e}")
            QtWidgets.QMessageBox.critical(
                self,
                "错误",
                "无法加载关于对话框模块"
            )
        except Exception as e:
            self.logger.error(f"显示关于对话框失败: {e}")
            QtWidgets.QMessageBox.critical(
                self,
                "错误", 
                f"无法显示关于对话框: {str(e)}"
            )

    def _save_epg_config(self, epg_config):
        """保存EPG配置回调"""
        try:
            if self.config.save_epg_config(epg_config):
                self.logger.info("EPG配置保存成功")
                # 重新初始化EPG管理器
                self.epg_manager = EPGManager(self.config, self)
                return True
            else:
                self.logger.warning("EPG配置保存失败")
                return False
        except Exception as e:
            self.logger.error(f"保存EPG配置出错: {e}")
            return False

    def save_before_exit(self):
        """程序退出前保存所有配置"""
        try:
            # 停止EPGManager线程
            if hasattr(self, 'epg_manager') and self.epg_manager.isRunning():
                self.epg_manager.quit()
                if not self.epg_manager.wait(2000):  # 等待2秒让线程退出
                    self.epg_manager.terminate()
                
            # 停止EPG UI线程
            if hasattr(self, 'epg_thread') and self.epg_thread.isRunning():
                self.epg_thread.quit()
                if not self.epg_thread.wait(1000):  # 等待1秒让线程退出
                    self.epg_thread.terminate()
            
            # 保存窗口布局
            size = self.size()
            dividers = [
                *self.ui.main_window.main_splitter.sizes(),
                *self.ui.main_window.left_splitter.sizes(),
                *self.ui.main_window.right_splitter.sizes(),
                *self.ui.main_window.h_splitter.sizes()
            ]
            self.config.save_window_layout(size.width(), size.height(), dividers)
            
            # 保存网络设置
            self.config.save_network_settings(
                self.ui.main_window.ip_range_input.text(),
                self.ui.main_window.timeout_input.value(),
                self.ui.main_window.thread_count_input.value(),
                self.ui.main_window.user_agent_input.text(),
                self.ui.main_window.referer_input.text()
            )
            self.logger.info("程序退出前配置已保存")
        except Exception as e:
            self.logger.error(f"保存退出配置失败: {e}")

def main():
    # 设置事件循环策略为Windows策略(兼容性更好)
    if sys.platform == "win32":
        import asyncio
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    
    # 创建应用实例
    app = QtWidgets.QApplication(sys.argv)
    
    # 创建主窗口
    window = MainWindow()
    window.show()
    
    # 确保程序退出前保存所有配置
    app.aboutToQuit.connect(window.save_before_exit)
    
    # 启动事件循环
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
